import openpyxl,os,sys
from openpyxl.utils import get_column_letter, column_index_from_string
# 进入同级目录下的file文件夹
currentFilePath = sys.path[0]
webFilePath = os.path.join(currentFilePath, "doc")
os.chdir(webFilePath)

wb = openpyxl.load_workbook('python.xlsx')
# 获取所有表名的列表
print(wb.get_sheet_names())
# 根据名字获取列表
sheet = wb.get_sheet_by_name('Yunis000')

print(sheet.title)
print(sheet['A1'])
print(sheet['A1'].value)

# A1 单元格
c = sheet['A1']
print('Row ' + str(c.row) + ', Column ' + str(c.column) + ' is ' + str(c.value))
# B10 单元格
c2 = sheet.cell(row=10, column=2)

print('Row ' + str(c2.row) + ', Column ' + str(c2.column) + ' is ' + str(c2.value)+"   __   "+str(c2.coordinate))

activeSheet = wb.get_active_sheet()
print("activeSheet " + activeSheet.title)

# 列字母和数字之间的转换
print(get_column_letter(1)) # A
print(column_index_from_string('A')) # 1




print(tuple(sheet['A1':'C3'])) #((<Cell 'Yunis000'.A1>, <Cell 'Yunis000'.B1>, <Cell 'Yunis000'.C1>), (<Cell 'Yunis000'.A2>, <Cell 'Yunis000'.B2>, <Cell 'Yunis000'.C2>), (<Cell 'Yunis000'.A3>, <Cell 'Yunis000'.B3>, <Cell 'Yunis000'.C3>))

for rowOfCellObjects in sheet['A1':'C3']:
    for cellObj in rowOfCellObjects:
        print(type(cellObj),type(rowOfCellObjects))
        print(cellObj.coordinate, cellObj.value)
    print('--- END OF ROW ---')

# A1 111111111
# B1 None
# C1 None
# --- END OF ROW ---
# A2 None
# B2 None
# C2 None
# --- END OF ROW ---
# A3 None
# B3 None
# C3 None
# --- END OF ROW ---
